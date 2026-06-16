# -*- coding: utf-8 -*-
"""
Genera el Excel editable del calendario a partir de datos/eventos.json.
Uso (una sola vez, o para regenerar el Excel desde el JSON):
    python datos/_generar_excel.py
"""
import json
import os
from datetime import date

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

AQUI = os.path.dirname(os.path.abspath(__file__))
JSON_PATH = os.path.join(AQUI, "eventos.json")
XLSX_PATH = os.path.join(AQUI, "calendario-el-vive.xlsx")

ANIO = 2026
ORDEN_MESES = ["Junio", "Julio", "Agosto", "Septiembre",
               "Octubre", "Noviembre", "Diciembre"]
MES_NUM = {m: i for i, m in enumerate(
    ["Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio", "Julio",
     "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre"], start=1)}

# Categoría interna (clave) -> nombre amigable que se muestra en el Excel
CAT_NOMBRE = {
    "comunidad": "Lunes de Junta",
    "apostolado": "Apostolado",
    "misa": "Misa",
    "matrimonios": "Matrimonios · KIDS · Juntas",
    "economica": "Económica",
    "especial": "Especial",
}

DOW_ES = ["Lun", "Mar", "Mié", "Jue", "Vie", "Sáb", "Dom"]  # date.weekday(): Lun=0

COLUMNS = ["Mes", "Día", "Día semana", "Categoría", "Título",
           "Hora", "Descripción", "Rango (varios días)"]


def dow_es(mes, dia):
    """Día de la semana en español para un día único; '' si no se puede calcular."""
    try:
        return DOW_ES[date(ANIO, MES_NUM[mes], int(dia)).weekday()]
    except (ValueError, KeyError, TypeError):
        return ""


def main():
    with open(JSON_PATH, encoding="utf-8") as f:
        eventos = json.load(f)

    # Ordena por mes (según ORDEN_MESES) y por día inicial
    def clave(e):
        try:
            d = int(str(e["dia"]).split("–")[0].split("-")[0].strip())
        except ValueError:
            d = 0
        return (ORDEN_MESES.index(e["mes"]) if e["mes"] in ORDEN_MESES else 99, d)

    eventos.sort(key=clave)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Calendario"

    # --- Estilos ---
    vino = "6E1E3C"
    head_fill = PatternFill("solid", fgColor=vino)
    head_font = Font(color="FFFFFF", bold=True, size=11)
    borde = Border(*[Side(style="thin", color="E2DAD6")] * 4)
    centro = Alignment(horizontal="center", vertical="center")
    izq = Alignment(horizontal="left", vertical="center", wrap_text=True)

    # --- Encabezado ---
    for c, titulo in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=c, value=titulo)
        cell.fill = head_fill
        cell.font = head_font
        cell.alignment = centro
        cell.border = borde

    # --- Filas ---
    for r, e in enumerate(eventos, start=2):
        rango = bool(e.get("rango"))
        dow = e.get("dow", "") if rango else (dow_es(e["mes"], e["dia"]) or e.get("dow", ""))
        valores = [
            e["mes"],
            e["dia"],
            dow,
            CAT_NOMBRE.get(e["cat"], e["cat"]),
            e["titulo"],
            e.get("hora", ""),
            e.get("desc", ""),
            "Sí" if rango else "",
        ]
        for c, v in enumerate(valores, start=1):
            cell = ws.cell(row=r, column=c, value=v)
            cell.border = borde
            cell.alignment = izq if c in (5, 7) else centro

    nfilas = len(eventos) + 1

    # --- Anchos de columna ---
    anchos = [13, 9, 11, 26, 48, 12, 50, 16]
    for i, w in enumerate(anchos, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 22

    # --- Validaciones (menús desplegables) ---
    dv_mes = DataValidation(type="list", formula1='"%s"' % ",".join(ORDEN_MESES), allow_blank=False)
    dv_cat = DataValidation(type="list", formula1='"%s"' % ",".join(CAT_NOMBRE.values()), allow_blank=False)
    dv_rango = DataValidation(type="list", formula1='"Sí,No"', allow_blank=True)
    ws.add_data_validation(dv_mes)
    ws.add_data_validation(dv_cat)
    ws.add_data_validation(dv_rango)
    dv_mes.add(f"A2:A{nfilas}")
    dv_cat.add(f"D2:D{nfilas}")
    dv_rango.add(f"H2:H{nfilas}")

    # --- Hoja de instrucciones ---
    ins = wb.create_sheet("Instrucciones")
    ins.column_dimensions["A"].width = 100
    lineas = [
        ("CALENDARIO ÉL VIVE · Cómo editar", True),
        ("", False),
        ("1. Edita esta hoja 'Calendario'. Cada fila es una actividad.", False),
        ("2. Para cambiar una fecha: modifica solo la columna 'Día' (y 'Mes' si cambia de mes).", False),
        ("   El 'Día semana' se recalcula automáticamente al actualizar; no necesitas tocarlo.", False),
        ("3. 'Categoría' y 'Mes' tienen menú desplegable. Usa siempre una de las opciones.", False),
        ("4. 'Hora' y 'Descripción' son opcionales (déjalas vacías si no aplican).", False),
        ("5. 'Rango (varios días)': pon 'Sí' solo en eventos que abarcan varios días", False),
        ("   (ej. Vacaciones). En esos casos escribe el rango en 'Día', ej. 3 – 13,", False),
        ("   y el 'Día semana' a mano, ej. 'Lun a Jue'.", False),
        ("6. Para AGREGAR un evento: añade una fila nueva. Para ELIMINAR: borra la fila.", False),
        ("7. Guarda el archivo. Luego pide: 'actualiza el calendario' y se genera el JSON", False),
        ("   y se actualiza la página web automáticamente.", False),
        ("", False),
        (f"Año del calendario: {ANIO}", False),
    ]
    for i, (txt, bold) in enumerate(lineas, start=1):
        cell = ins.cell(row=i, column=1, value=txt)
        cell.font = Font(bold=bold, size=14 if bold else 11, color=vino if bold else "333333")
        cell.alignment = Alignment(wrap_text=True, vertical="top")

    wb.save(XLSX_PATH)
    print("Excel generado:", XLSX_PATH, "| filas:", len(eventos))


if __name__ == "__main__":
    main()
